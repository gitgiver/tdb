import os
import re
import logging
from typing import List, Dict, Any, Optional, Tuple
import numpy as np
from tqdm import tqdm
import faiss
import pickle
import torch
from sentence_transformers import SentenceTransformer

from langchain.text_splitter import RecursiveCharacterTextSplitter
from rank_bm25 import BM25Okapi
import jieba
import nltk
import argparse
import json
import time
import requests
import zhipuai  # 智谱官方SDK
import pandas as pd
from abc import ABC, abstractmethod
from pathlib import Path


# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class TextProcessor:
    """文本处理类：负责读取文本文件并进行预处理"""

    def __init__(self, data_dir: str):
        """
        初始化文本处理器

        Args:
            data_dir: 文本文件所在目录
        """
        self.data_dir = data_dir
        self.txt_file_paths = []
        self.excel_file_paths = []
        self._load_file_paths()

    def _load_file_paths(self) -> None:
        """加载目录中的所有txt文件和xlsx文件路径"""
        if not self.data_dir or not os.path.exists(self.data_dir):
            logger.warning(f"目录不存在或为空: {self.data_dir}")
            return

        self.txt_file_paths = []
        self.excel_file_paths = []

        for f in os.listdir(self.data_dir):
            full_path = os.path.join(self.data_dir, f)
            if f.endswith('.txt'):
                self.txt_file_paths.append(full_path)
            elif f.endswith('.xlsx'):
                self.excel_file_paths.append(full_path)

        logger.info(f"找到 {len(self.txt_file_paths)} 个txt文件和 {len(self.excel_file_paths)} 个Excel文件")

    def read_text_file(self, file_path: str) -> str:
        """
        读取文本文件内容，并在文本开头添加文件名作为关键标识

        Args:
            file_path: 文件路径

        Returns:
            带有文件名前缀的文件内容
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()

            # 提取文件名并添加到文本开头
            file_name = os.path.basename(file_path)
            # 移除扩展名
            file_name_without_ext = os.path.splitext(file_name)[0]

            # 格式化文件名，提取关键信息
            # 例如：将"01_智能数据采集装置设计专项赛.txt"转为"智能数据采集装置设计专项赛"
            cleaned_name = re.sub(r'^\d+[_\-]', '', file_name_without_ext)

            # 将文件名添加到文本开头，作为最重要的上下文
            enhanced_content = f"文档标题: {cleaned_name}\n\n{content}"
            return enhanced_content

        except Exception as e:
            logger.error(f"读取文件 {file_path} 时出错: {e}")
            return ""

    def clean_text(self, text: str) -> str:
        """
        清理文本

        Args:
            text: 原始文本

        Returns:
            清理后的文本
        """
        # 删除多余的空白字符
        text = re.sub(r'\s+', ' ', text)
        # 删除可能的特殊字符
        text = re.sub(r'[^\w\s\u4e00-\u9fff.,，。、:：?？!！;；]', '', text)
        return text.strip()

    def get_document_metadata(self, file_path: str) -> Dict[str, Any]:
        """
        获取文档元数据

        Args:
            file_path: 文件路径

        Returns:
            文档元数据
        """
        file_name = os.path.basename(file_path)
        return {
            "source": file_name,
            "file_path": file_path,
            "created_at": os.path.getctime(file_path)
        }

    def read_excel_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        读取Excel文件并转换为文本列表

        Args:
            file_path: Excel文件路径

        Returns:
            包含文本和元数据的字典列表
        """
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path)
            logger.info(f"读取Excel文件: {file_path}, 包含 {len(df)} 行数据")

            # 准备结果列表
            result = []

            # 处理每一行数据
            for index, row in df.iterrows():
                # 将行数据转换为字典
                row_dict = row.to_dict()

                # 创建文本内容 - 将所有列的内容合并为一个文本
                text_parts = []
                for col_name, value in row_dict.items():
                    # 跳过空值
                    if pd.isna(value):
                        continue
                    # 添加列名和值
                    text_parts.append(f"{col_name}: {value}")

                # 合并所有部分为一个文本
                text = "\n".join(text_parts)

                # 创建元数据
                metadata = {
                    "source": f"Excel行{index+1}",
                    "file_path": file_path,
                    "row_index": index,
                    "created_at": os.path.getctime(file_path)
                }

                # 添加到结果列表
                if text.strip():  # 确保文本不为空
                    result.append({"text": text, "metadata": metadata})

            logger.info(f"从Excel文件中提取了 {len(result)} 条有效文本记录")
            return result

        except Exception as e:
            logger.error(f"读取Excel文件 {file_path} 时出错: {e}")
            return []


class TextSplitter:
    """文本分块类：负责将文本分割成适合向量化的块"""

    def __init__(self,
                chunk_size: int = 500,
                chunk_overlap: int = 50):
        """
        初始化文本分块器

        Args:
            chunk_size: 块大小（字符数）
            chunk_overlap: 块重叠大小（字符数）
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1200,      # 增大分块大小
            chunk_overlap=250,    # 增加重叠部分
            separators=["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""]
        )

    def split_text(self, text: str) -> List[str]:
        """
        分割文本

        Args:
            text: 需要分割的文本

        Returns:
            分割后的文本块列表
        """
        if not text:
            return []

        try:
            chunks = self.text_splitter.split_text(text)
            logger.info(f"文本被分割成 {len(chunks)} 个块")
            return chunks
        except Exception as e:
            logger.error(f"分割文本时出错: {e}")
            return []

    def split_documents(self, texts: List[str], metadatas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        分割多个文档并保留元数据

        Args:
            texts: 文档文本列表
            metadatas: 对应的元数据列表

        Returns:
            带有文本和元数据的分块列表
        """
        documents = []

        for i, (text, metadata) in enumerate(zip(texts, metadatas)):
            chunks = self.split_text(text)
            for j, chunk in enumerate(chunks):
                chunk_metadata = metadata.copy()
                chunk_metadata.update({
                    "chunk_id": j,
                    "total_chunks": len(chunks)
                })
                documents.append({
                    "text": chunk,
                    "metadata": chunk_metadata
                })

        return documents


class TextVectorizer:
    """文本向量化类：负责将文本转换为向量表示"""

    def __init__(self, model_name: str = "all-MiniLM-L6-v2"):
        """
        初始化文本向量化器

        Args:
            model_name: 向量化模型名称
        """
        logger.info(f"加载向量化模型: {model_name}")
        # self.device = "cuda" if torch.cuda.is_available() else "cpu"
        self.device="cpu"
        logger.info(f"使用设备: {self.device}")

        # 使用to_empty()替代to()避免meta tensor错误
        if self.device == "cuda":
            self.model = SentenceTransformer(model_name).to_empty(device=self.device)
        else:
            self.model = SentenceTransformer(model_name, device=self.device)

        self.vector_dim = self.model.get_sentence_embedding_dimension()
        logger.info(f"向量维度: {self.vector_dim}")

    def vectorize(self, texts: List[str], batch_size: int = 32) -> np.ndarray:
        """
        将文本列表转换为向量表示

        Args:
            texts: 文本列表
            batch_size: 批处理大小

        Returns:
            文本向量数组
        """
        if not texts:
            return np.array([])

        try:
            logger.info(f"开始向量化 {len(texts)} 个文本块")
            embeddings = self.model.encode(
                texts,
                batch_size=batch_size,
                show_progress_bar=True,
                convert_to_numpy=True
            )
            return embeddings
        except Exception as e:
            logger.error(f"向量化文本时出错: {e}")
            return np.array([])


class VectorDatabase:
    """向量数据库类：负责存储和检索向量数据"""

    def __init__(self, vector_dim: int):
        """
        初始化向量数据库

        Args:
            vector_dim: 向量维度
        """
        self.vector_dim = vector_dim
        self.index = faiss.IndexFlatL2(vector_dim)  # L2距离索引
        self.documents = []  # 存储文档及其元数据

    def add_documents(self, vectors: np.ndarray, documents: List[Dict[str, Any]]) -> None:
        """
        添加文档向量和元数据到数据库

        Args:
            vectors: 文档向量
            documents: 文档及其元数据
        """
        if vectors.size == 0 or not documents:
            logger.warning("没有向量或文档可添加")
            return

        try:
            # 确保向量为float32类型
            vectors = vectors.astype(np.float32)
            # 添加向量到索引
            self.index.add(vectors)
            # 存储文档
            start_idx = len(self.documents)
            for i, doc in enumerate(documents):
                doc["vector_id"] = start_idx + i
                self.documents.append(doc)

            logger.info(f"添加了 {len(documents)} 个文档到向量数据库")
        except Exception as e:
            logger.error(f"添加文档到向量数据库时出错: {e}")

    def search(self, query_vector: np.ndarray, top_k: int = 10) -> List[Dict[str, Any]]:
        """
        搜索最相似的文档

        Args:
            query_vector: 查询向量
            top_k: 返回的结果数量

        Returns:
            相似文档列表
        """
        if self.index.ntotal == 0:
            logger.warning("向量数据库为空，无法搜索")
            return []

        # 确保查询向量为正确形状的float32数组
        query_vector = query_vector.reshape(1, -1).astype(np.float32)

        # 执行搜索
        distances, indices = self.index.search(query_vector, min(top_k, self.index.ntotal))

        # 获取匹配的文档
        results = []
        for i, idx in enumerate(indices[0]):
            if idx < len(self.documents):
                doc = self.documents[idx].copy()
                doc["distance"] = float(distances[0][i])
                results.append(doc)

        return results

    def save(self, file_path: str) -> None:
        """
        保存向量数据库到文件

        Args:
            file_path: 保存路径
        """
        try:
            data = {
                "vector_dim": self.vector_dim,
                "documents": self.documents
            }

            # 保存文档数据
            with open(file_path + ".documents.pkl", "wb") as f:
                pickle.dump(data, f)

            # 保存向量索引
            faiss.write_index(self.index, file_path + ".index")

            logger.info(f"向量数据库已保存到 {file_path}")
        except Exception as e:
            logger.error(f"保存向量数据库时出错: {e}")

    @classmethod
    def load(cls, file_path: str) -> "VectorDatabase":
        """
        从文件加载向量数据库

        Args:
            file_path: 加载路径

        Returns:
            加载的向量数据库
        """
        try:
            # 加载文档数据
            with open(file_path + ".documents.pkl", "rb") as f:
                data = pickle.load(f)

            vector_dim = data["vector_dim"]
            documents = data["documents"]

            # 创建数据库实例
            db = cls(vector_dim)
            db.documents = documents

            # 加载向量索引
            db.index = faiss.read_index(file_path + ".index")

            logger.info(f"已从 {file_path} 加载向量数据库，包含 {len(documents)} 个文档")
            return db
        except Exception as e:
            logger.error(f"加载向量数据库时出错: {e}")
            return cls(0)  # 创建空数据库


class BM25Retriever:
    """BM25稀疏检索器：基于关键词匹配的检索方法"""

    def __init__(self):
        """初始化BM25检索器"""
        self.bm25 = None
        self.tokenized_corpus = []
        self.documents = []

        # 确保nltk的punkt分词器已下载
        try:
            nltk.data.find('tokenizers/punkt')
        except LookupError:
            nltk.download('punkt')

    def _tokenize(self, text: str) -> List[str]:
        """分词函数，支持中英文混合分词"""
        # 检测文本中是否包含中文
        if any('\u4e00' <= char <= '\u9fff' for char in text):
            # 使用jieba进行中文分词
            return list(jieba.cut(text))
        else:
            # 使用nltk进行英文分词
            return nltk.word_tokenize(text.lower())

    def fit(self, texts: List[str], documents: List[Dict[str, Any]]) -> None:
        """
        训练BM25模型

        Args:
            texts: 文档文本列表
            documents: 文档及其元数据列表
        """
        if not texts or not documents:
            logger.warning("没有文档可供BM25模型训练")
            return

        logger.info(f"开始为BM25模型分词处理 {len(texts)} 个文档...")

        # 对所有文档进行分词
        self.tokenized_corpus = [self._tokenize(text) for text in texts]

        # 创建BM25模型
        self.bm25 = BM25Okapi(self.tokenized_corpus)

        # 存储文档引用
        self.documents = documents

        logger.info("BM25模型训练完成")

    def search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """
        使用BM25算法搜索相关文档

        Args:
            query: 查询文本
            top_k: 返回结果数量

        Returns:
            相似文档列表
        """
        if not self.bm25:
            logger.warning("BM25模型未训练，无法执行搜索")
            return []

        # 对查询文本进行分词
        tokenized_query = self._tokenize(query)

        # 执行BM25搜索
        scores = self.bm25.get_scores(tokenized_query)

        # 获取top_k的索引和分数
        top_indices = np.argsort(scores)[::-1][:top_k]

        # 构建结果
        results = []
        for idx in top_indices:
            if scores[idx] > 0:  # 只返回有匹配的结果
                doc = self.documents[idx].copy()
                doc["bm25_score"] = float(scores[idx])
                results.append(doc)

        return results


class HybridRetriever:
    """混合检索系统，结合稠密检索和稀疏检索"""

    def __init__(self, dense_retriever, sparse_retriever, dense_weight: float = 0.7):#####试试改改这个
        """
        初始化混合检索系统

        Args:
            dense_retriever: 稠密检索器
            sparse_retriever: 稀疏检索器
            dense_weight: 稠密检索权重 (0.5，平衡两种检索方式)
        """
        self.dense_retriever = dense_retriever
        self.sparse_retriever = sparse_retriever
        self.dense_weight = dense_weight
        logger.info(f"混合检索系统初始化完成，稠密检索权重: {dense_weight}")

    def search(self, query: str, top_k: int = 10, threshold: float = 0.5) -> List[Dict]:
        """
        执行混合检索

        Args:
            query: 查询文本
            top_k: 返回的结果数量
            threshold: 匹配度阈值，小于该值的结果将被过滤

        Returns:
            检索结果列表，每个结果包含文档ID、文本内容、匹配分数等信息
        """
        try:
            # 1. 先提取查询中可能的竞赛名称
            competition_name = None
            competition_match = re.search(r'[""]?([^""]+(?:专项赛|挑战赛|竞赛|大赛))[""]?', query)
            if competition_match:
                competition_name = competition_match.group(1)
                logger.info(f"从查询中提取到竞赛名称: {competition_name}")

            # 2. 执行稠密检索和稀疏检索
            dense_results = self.dense_retriever.search(query, top_k=top_k*2)
            sparse_results = self.sparse_retriever.search(query, top_k=top_k*2)

            # 3. 合并结果
            all_doc_ids = set([r['id'] for r in dense_results] + [r['id'] for r in sparse_results])
            merged_results = {}

            # 计算合并分数
            for doc_id in all_doc_ids:
                # 查找文档在两种检索结果中的分数
                dense_score = next((r['score'] for r in dense_results if r['id'] == doc_id), 0)
                sparse_score = next((r['score'] for r in sparse_results if r['id'] == doc_id), 0)

                # 计算混合分数，稠密检索和稀疏检索的加权平均
                hybrid_score = self.dense_weight * dense_score + (1 - self.dense_weight) * sparse_score

                # 如果提取到竞赛名称，检查文档内容是否包含该名称，提高优先级
                if competition_name:
                    doc_content = next((r['content'] for r in dense_results if r['id'] == doc_id), "")
                    if not doc_content:
                        doc_content = next((r['content'] for r in sparse_results if r['id'] == doc_id), "")

                    # 如果文档内容包含竞赛名称，大幅提高其分数
                    if competition_name in doc_content:
                        # 加权提升，确保竞赛名称匹配的文档优先级最高
                        hybrid_score *= 1.5

                # 判断是否达到阈值
                if hybrid_score >= threshold:
                    # 获取文档内容
                    doc_content = next((r['content'] for r in dense_results if r['id'] == doc_id), None)
                    if not doc_content:
                        doc_content = next((r['content'] for r in sparse_results if r['id'] == doc_id), None)

                    # 获取文档元数据
                    doc_metadata = next((r.get('metadata', {}) for r in dense_results if r['id'] == doc_id), None)
                    if not doc_metadata:
                        doc_metadata = next((r.get('metadata', {}) for r in sparse_results if r['id'] == doc_id), None)

                    # 精确关键词匹配评分
                    keyword_match_score = self._calculate_keyword_match_score(query, doc_content)

                    # 组合最终分数：混合分数 + 关键词匹配加成
                    final_score = hybrid_score + keyword_match_score * 0.3

                    merged_results[doc_id] = {
                        'id': doc_id,
                        'content': doc_content,
                        'score': final_score,
                        'metadata': doc_metadata
                    }

            # 4. 根据最终分数排序并返回前K个结果
            sorted_results = sorted(merged_results.values(), key=lambda x: x['score'], reverse=True)
            return sorted_results[:top_k]

        except Exception as e:
            logger.error(f"混合检索过程中出错: {e}")
            return []

    def _calculate_keyword_match_score(self, query: str, content: str) -> float:
        """
        计算查询与文档内容的关键词匹配分数

        Args:
            query: 查询文本
            content: 文档内容

        Returns:
            关键词匹配分数 (0-1)
        """
        if not content:
            return 0

        # 提取查询中的关键词
        query_keywords = jieba.lcut_for_search(query)
        # 过滤停用词和短词
        query_keywords = [k for k in query_keywords if len(k) > 1]

        if not query_keywords:
            return 0

        # 计算匹配的关键词数量
        matched_keywords = [k for k in query_keywords if k in content]
        match_ratio = len(matched_keywords) / len(query_keywords)

        # 特别加权匹配的特殊关键词（如竞赛名称、时间等）
        special_patterns = [
            r'报名时间', r'比赛时间', r'组队要求', r'参赛要求',
            r'官网', r'官方网站', r'专项赛', r'挑战赛'
        ]

        special_bonus = 0
        for pattern in special_patterns:
            if re.search(pattern, query) and re.search(pattern, content):
                special_bonus += 0.1  # 特殊关键词匹配加分

        return min(1.0, match_ratio + special_bonus)  # 最高分为1


class RetrievalSystem:
    """检索系统，集成向量检索和关键词检索功能，增强了语义理解和结构感知能力"""

    def __init__(self,
                 vector_model_name: str = "shibing624/text2vec-base-chinese",
                 persist_dir: Optional[str] = None):
        """
        初始化检索系统

        Args:
            vector_model_name: 向量化模型名称
            persist_dir: 索引持久化目录，如果指定，则尝试从此目录加载索引
        """
        self.vector_model_name = vector_model_name
        self.persist_dir = persist_dir
        self.index_built = False
        self.docs = []

        # 初始化关键词增强器
        self.keyword_enhancer = KeywordEnhancer()

        try:
            # 尝试导入必要的库
            from sentence_transformers import SentenceTransformer
            import faiss

            # 加载向量模型
            logger.info(f"加载向量模型: {vector_model_name}")


            device = torch.device('cpu')  # 强制使用CPU本来就无gpu
            self.vector_model = SentenceTransformer(vector_model_name, device=device)
            self.vector_dim = self.vector_model.get_sentence_embedding_dimension()

            # 初始化FAISS索引
            self.vector_index = faiss.IndexFlatIP(self.vector_dim)  # 内积相似度(余弦相似度)

            # 初始化BM25索引
            self._init_bm25_index()

            # 如果指定了持久化目录，尝试加载索引
            if persist_dir and os.path.exists(persist_dir):
                self.load_index(persist_dir)

        except ImportError as e:
            logger.error(f"导入必要库失败: {str(e)}")
            raise e

    def _init_bm25_index(self):
        """初始化BM25索引"""
        try:
            from rank_bm25 import BM25Okapi
            self.bm25_class = BM25Okapi
            self.bm25_index = None
            self.tokenized_corpus = []
        except ImportError:
            logger.warning("无法导入rank_bm25库，BM25检索将不可用")
            self.bm25_class = None

    def _tokenize_text(self, text: str) -> List[str]:
        """
        分词函数

        Args:
            text: 待分词的文本

        Returns:
            分词结果列表
        """
        # 使用jieba分词
        return list(jieba.cut(text))

    def build_index(self, documents: List[Dict[str, Any]]) -> None:
        """
        构建索引

        Args:
            documents: 文档列表，每个文档应包含text和metadata字段
        """
        if not documents:
            logger.warning("没有文档可供索引，索引构建失败")
            return

        # 存储文档
        self.docs = documents

        # 提取文本和元数据
        texts = [doc["text"] for doc in documents]

        # 向量化文本
        logger.info(f"向量化{len(texts)}个文档...")
        embeddings = self.vector_model.encode(texts, show_progress_bar=True)

        # 转换为适合FAISS的格式
        embeddings = np.array(embeddings).astype('float32')

        # 重置并构建FAISS索引
        self.vector_index = faiss.IndexFlatIP(self.vector_dim)
        self.vector_index.add(embeddings)

        # 构建BM25索引
        if self.bm25_class:
            logger.info("构建BM25索引...")
            # 对所有文档进行分词
            self.tokenized_corpus = [self._tokenize_text(text) for text in texts]
            # 建立BM25索引
            self.bm25_index = self.bm25_class(self.tokenized_corpus)

        self.index_built = True
        logger.info(f"索引构建完成，包含{len(documents)}个文档")

        # 如果指定了持久化目录，保存索引
        if self.persist_dir:
            self.save_index(self.persist_dir)

    def search(self,
               query: str,
               top_k: int = 10,
               hybrid_weight: float = 0.7) -> List[Dict[str, Any]]:
        """
        混合检索

        Args:
            query: 查询文本
            top_k: 返回结果数量
            hybrid_weight: 向量检索权重，BM25权重为1-hybrid_weight

        Returns:
            检索结果列表
        """
        if not self.index_built or not self.docs:
            logger.warning("索引未构建或没有文档，无法执行检索")
            return []

        # 增强查询，添加同义词和结构信息
        enhanced_query = self.keyword_enhancer.enhance_query(query)
        logger.info(f"增强后的查询: {enhanced_query}")

        # 1. 先进行精确匹配搜索
        exact_match_results = self._exact_match_search(query, top_k=top_k)

        # 2. 向量检索
        vector_results = self._vector_search(enhanced_query, top_k=top_k*2)  # 检索更多结果用于混合

        # 3. 关键词检索 (如果BM25可用)
        keyword_results = []
        if self.bm25_index:
            keyword_results = self._keyword_search(enhanced_query, top_k=top_k*2)

        # 4. 混合排序
        if exact_match_results:
            # 如果有精确匹配，优先返回
            logger.info(f"找到{len(exact_match_results)}个精确匹配结果")
            if len(exact_match_results) >= top_k:
                return exact_match_results[:top_k]
            else:
                # 将剩余位置用混合结果填充
                remaining = top_k - len(exact_match_results)
                hybrid_results = self._hybrid_rank(vector_results, keyword_results, hybrid_weight)

                # 去除已经在精确匹配中的文档
                exact_match_ids = [result["id"] for result in exact_match_results]
                filtered_hybrid = [r for r in hybrid_results if r["id"] not in exact_match_ids]

                return exact_match_results + filtered_hybrid[:remaining]
        else:
            # 如果没有精确匹配，返回混合排序结果
            return self._hybrid_rank(vector_results, keyword_results, hybrid_weight)[:top_k]

    def _vector_search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """
        向量检索

        Args:
            query: 查询文本
            top_k: 返回结果数量

        Returns:
            检索结果列表
        """
        # 提取竞赛名称
        competition_name = None
        competition_match = re.search(r'[""]?([^""]+?(?:专项赛|挑战赛|竞赛|大赛))[""]?', query)
        if competition_match:
            competition_name = competition_match.group(1)

        # 提取关键查询目标
        is_prep_query = any(term in query for term in ["准备", "准备工作", "参赛准备", "报名准备"])

        # 向量化查询
        query_vector = self.vector_model.encode([query])[0].astype('float32')
        query_vector = query_vector.reshape(1, -1)

        # 执行检索
        scores, indices = self.vector_index.search(query_vector, min(top_k * 3, self.vector_index.ntotal))  # 检索更多结果，便于后处理

        # 格式化结果
        results = []
        for i, (score, idx) in enumerate(zip(scores[0], indices[0])):
            if idx < 0 or idx >= len(self.docs):
                continue  # 跳过无效索引

            doc = self.docs[idx]
            text = doc["text"]
            metadata = doc.get("metadata", {})
            source = metadata.get("source", "")

            # 获取文档结构信息
            section = metadata.get("section", "")
            key_info = metadata.get("key_info", {})

            # 计算额外加分
            structure_bonus = 0
            if section:
                # 检查章节是否与查询相关
                for keyword in self.keyword_enhancer.extract_keywords(query):
                    if keyword in section:
                        structure_bonus += 0.1  # 章节匹配加分

            # 检查是否包含关键信息
            for key, value in key_info.items():
                if key in query or any(kw in key for kw in self.keyword_enhancer.extract_keywords(query)):
                    structure_bonus += 0.2  # 包含查询相关的关键信息加分

            # 特殊处理未来校园智能应用专项赛
            competition_bonus = 0
            if competition_name:
                clean_comp = competition_name.replace('"', '').replace('"', '')

                # 未来校园智能应用专项赛特殊处理
                if "未来校园" in clean_comp:
                    if "未来校园" in source or "未来校园" in text:
                        competition_bonus += 1.0
                        # 如果查询涉及准备工作，检查文档中是否包含相关段落
                        if is_prep_query and any(prep_term in text for prep_term in ["参赛要求", "报名方式", "参赛对象", "流程", "准备"]):
                            competition_bonus += 2.0  # 准备工作相关内容特别加分

            # 调整最终分数
            final_score = float(score) + structure_bonus + competition_bonus

            results.append({
                "id": idx,
                "text": doc["text"],
                "metadata": doc["metadata"],
                "score": final_score,
                "vector_score": float(score),
                "structure_bonus": structure_bonus,
                "competition_bonus": competition_bonus,
                "rank": i
            })

        # 后处理提升相关性
        if is_prep_query and competition_name and "未来校园" in competition_name:
            # 特别检查所有文档中是否包含"未来校园"
            for doc in self.docs:
                idx = self.docs.index(doc)
                if idx not in [r["id"] for r in results]:  # 如果文档不在结果中
                    source = doc.get("metadata", {}).get("source", "")
                    text = doc.get("text", "")

                    # 检查是否是未来校园专项赛文档
                    if ("未来校园" in source or "01_" in source) and "智能应用" in source:
                        # 查找准备工作相关内容
                        prep_score = 0
                        for prep_term in ["参赛要求", "报名方式", "参赛对象", "流程", "准备"]:
                            if prep_term in text:
                                prep_score += 0.5

                        if prep_score > 0:
                            # 添加到结果中
                            results.append({
                                "id": idx,
                                "text": doc["text"],
                                "metadata": doc["metadata"],
                                "score": 0.6 + prep_score,  # 基础分 + 准备工作相关度
                                "vector_score": 0.6,  # 基础向量分数
                                "structure_bonus": 0,
                                "competition_bonus": prep_score,
                                "rank": len(results),
                                "added_manually": True
                            })

        # 按分数排序
        results.sort(key=lambda x: x["score"], reverse=True)

        # 更新排名
        for i, result in enumerate(results):
            result["rank"] = i

        return results[:top_k]

    def _keyword_search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """
        关键词检索，已增强同义词扩展

        Args:
            query: 查询文本
            top_k: 返回结果数量

        Returns:
            检索结果列表
        """
        if not self.bm25_index:
            return []

        # 提取查询关键词
        keywords = self.keyword_enhancer.extract_keywords(query)

        # 扩展同义词
        expanded_keywords = []
        for keyword in keywords:
            synonyms = self.keyword_enhancer.get_synonyms(keyword)
            expanded_keywords.extend(synonyms)

        # 去重
        expanded_keywords = list(set(expanded_keywords))

        logger.info(f"扩展后的关键词: {expanded_keywords}")

        # 对查询分词
        tokenized_query = []
        for keyword in expanded_keywords:
            tokenized_query.extend(self._tokenize_text(keyword))

        # 执行BM25检索
        bm25_scores = self.bm25_index.get_scores(tokenized_query)

        # 获取前top_k结果
        top_indices = np.argsort(bm25_scores)[::-1][:top_k]

        # 格式化结果
        results = []
        for i, idx in enumerate(top_indices):
            score = bm25_scores[idx]
            if score <= 0:  # 跳过无关结果
                continue

            doc = self.docs[idx]

            # 计算关键词匹配度加成
            keyword_match_bonus = 0
            for keyword in expanded_keywords:
                if keyword in doc["text"]:
                    # 关键词在原始查询中权重更高
                    if keyword in keywords:
                        keyword_match_bonus += 0.1
                    else:  # 同义词权重稍低
                        keyword_match_bonus += 0.05

            # 调整最终分数
            final_score = float(score) + keyword_match_bonus

            results.append({
                "id": idx,
                "text": doc["text"],
                "metadata": doc["metadata"],
                "score": final_score,
                "keyword_score": float(score),
                "keyword_match_bonus": keyword_match_bonus,
                "rank": i
            })

        return results

    def _exact_match_search(self, query: str, top_k: int = 10) -> List[Dict[str, Any]]:
        """
        精确匹配搜索，优先匹配竞赛名称和任务编号

        Args:
            query: 查询文本
            top_k: 返回结果数量

        Returns:
            检索结果列表
        """
        # 对查询进行清理和分解
        query = query.strip()

        # 处理带引号的竞赛名称 - 改进匹配算法
        competition_match = re.search(r'[""]?([^""]+?(?:专项赛|挑战赛|竞赛|大赛))[""]?', query)
        competition_name = None
        comp_alternatives = []

        if competition_match:
            competition_name = competition_match.group(1)
            logger.info(f"从查询中提取到竞赛名称: {competition_name}")

            # 创建竞赛名称的替代形式
            # 1. 去掉引号版本
            clean_name = competition_name.replace('"', '').replace('"', '')
            comp_alternatives.append(clean_name)

            # 2. 加引号版本
            quoted_name = f'"{clean_name}"'
            comp_alternatives.append(quoted_name)

            # 3. 前缀处理 (例如针对"01_"未来校园"智能应用专项赛.txt"这种情况)
            if "未来校园" in clean_name:
                comp_alternatives.append("未来校园智能应用专项赛")
                comp_alternatives.append("未来校园")

            logger.info(f"竞赛名称替代形式: {comp_alternatives}")

        # 提取任务编号
        task_match = re.search(r'(任务[一二三四五六七八九十\d]+)', query)
        task_num = None
        if task_match:
            task_num = task_match.group(1)
            logger.info(f"从查询中提取到任务编号: {task_num}")

        # 准备关键词列表
        keywords = []
        if competition_name:
            keywords.append(competition_name)
            keywords.extend(comp_alternatives)
        if task_num:
            keywords.append(task_num)

        # 添加其他关键词
        other_keywords = self.keyword_enhancer.extract_keywords(query)
        for kw in other_keywords:
            if kw not in keywords:
                keywords.append(kw)

        if not keywords:
            return []

        logger.info(f"精确匹配关键词: {keywords}")

        # 对文档进行精确匹配检索
        results = []
        for idx, doc in enumerate(self.docs):
            text = doc["text"]
            metadata = doc.get("metadata", {})
            source = metadata.get("source", "")

            # 首先检查文件名是否匹配竞赛名称
            file_match_score = 0
            if competition_name:
                # 检查原始名称
                if competition_name in source:
                    file_match_score = 2.0  # 文件名匹配得高分
                else:
                    # 检查替代形式
                    for alt_name in comp_alternatives:
                        if alt_name in source:
                            file_match_score = 2.0  # 文件名匹配得高分
                            break

                    # 特殊处理未来校园专项赛
                    if "未来校园" in source and "智能应用" in source:
                        file_match_score = 2.5  # 特别加分

            # 检查是否包含关键信息
            key_info_score = 0
            key_info = metadata.get("key_info", {})
            for key, value in key_info.items():
                if key in query or any(kw in key for kw in keywords):
                    key_info_score += 0.5  # 关键信息匹配加分

            # 检查是否匹配章节
            section_score = 0
            section = metadata.get("section", "")
            if section:
                for kw in keywords:
                    if kw in section:
                        section_score += 0.5  # 章节匹配加分

            # 计算内容匹配分数
            content_match_count = 0

            # 竞赛名称匹配（高优先级）
            if competition_name:
                # 检查原始竞赛名称
                if competition_name in text:
                    content_match_count += 3  # 竞赛名称匹配权重更高
                else:
                    # 检查替代形式
                    for alt_name in comp_alternatives:
                        if alt_name in text:
                            content_match_count += 3  # 替代形式匹配也同样重要
                            break

                    # 特殊处理未来校园
                    if "未来校园" in text and "智能应用" in text:
                        content_match_count += 4  # 特别加分

            # 任务编号匹配（高优先级）
            if task_num:
                if task_num in text:
                    content_match_count += 3  # 任务编号匹配权重更高
                    # 检查任务描述是否在同一段
                    task_desc_pattern = fr"{task_num}[：:]\s*(.+?)(?:\n\n|$)"
                    if re.search(task_desc_pattern, text):
                        content_match_count += 2  # 包含任务描述加分

            # 其他关键词匹配
            for kw in keywords:
                if kw != competition_name and kw != task_num and kw in text:
                    content_match_count += 1

            # 计算总分
            total_score = file_match_score + key_info_score + section_score + (content_match_count / max(1, len(keywords)))

            # 准备工作匹配加分
            if "准备工作" in query and any(term in text for term in ["准备", "准备工作", "参赛准备", "报名准备"]):
                total_score += 1.5  # 准备工作相关内容加分

            if total_score > 0:
                # 添加到结果中
                results.append({
                    "id": idx,
                    "text": doc["text"],
                    "metadata": doc["metadata"],
                    "score": total_score,
                    "match_type": "exact_match",
                    "file_match_score": file_match_score,
                    "content_match_score": content_match_count / max(1, len(keywords)),
                    "key_info_score": key_info_score,
                    "section_score": section_score,
                    "rank": 0  # 会在后续排序
                })

        # 按分数排序
        results.sort(key=lambda x: x["score"], reverse=True)

        # 更新排名
        for i, result in enumerate(results):
            result["rank"] = i

        logger.info(f"精确匹配找到 {len(results)} 个结果")

        return results[:top_k]

    def _hybrid_rank(self,
                    vector_results: List[Dict[str, Any]],
                    keyword_results: List[Dict[str, Any]],
                    vector_weight: float = 0.7) -> List[Dict[str, Any]]:
        """
        混合排序

        Args:
            vector_results: 向量检索结果
            keyword_results: 关键词检索结果
            vector_weight: 向量检索权重

        Returns:
            混合排序结果
        """
        # 如果其中一种方法无结果，直接返回另一个方法的结果
        if not vector_results and not keyword_results:
            logger.warning("向量检索和关键词检索都没有返回结果")
            return []

        if not vector_results:
            logger.info("仅使用关键词检索结果")
            return keyword_results

        if not keyword_results:
            logger.info("仅使用向量检索结果")
            return vector_results

        # 合并结果并按文档ID去重
        doc_scores = {}

        # 处理向量检索结果
        max_vector_score = float('-inf')
        min_vector_score = float('inf')

        for result in vector_results:
            doc_id = result["id"]
            # 确保分数有效
            score = result.get("score", 0)
            if isinstance(score, (int, float)) and score > 0:
                max_vector_score = max(max_vector_score, score)
                min_vector_score = min(min_vector_score, score)

            if doc_id not in doc_scores:
                doc_scores[doc_id] = {
                    "doc": result,
                    "vector_score": 0,
                    "keyword_score": 0
                }
            doc_scores[doc_id]["vector_score"] = score

        # 处理关键词检索结果
        max_keyword_score = float('-inf')
        min_keyword_score = float('inf')

        for result in keyword_results:
            doc_id = result["id"]
            # 确保分数有效
            score = result.get("score", 0)
            if isinstance(score, (int, float)) and score > 0:
                max_keyword_score = max(max_keyword_score, score)
                min_keyword_score = min(min_keyword_score, score)

            if doc_id not in doc_scores:
                doc_scores[doc_id] = {
                    "doc": result,
                    "vector_score": 0,
                    "keyword_score": 0
                }
            doc_scores[doc_id]["keyword_score"] = score

        # 处理归一化边界值问题
        if max_vector_score == float('-inf'):
            max_vector_score = 1
        if min_vector_score == float('inf'):
            min_vector_score = 0

        if max_keyword_score == float('-inf'):
            max_keyword_score = 1
        if min_keyword_score == float('inf'):
            min_keyword_score = 0

        # 计算混合分数
        results = []
        for doc_id, scores in doc_scores.items():
            # 归一化向量分数
            if max_vector_score > min_vector_score:
                normalized_vector_score = (scores["vector_score"] - min_vector_score) / (max_vector_score - min_vector_score)
            else:
                normalized_vector_score = 1.0 if scores["vector_score"] > 0 else 0.0

            # 归一化关键词分数
            if max_keyword_score > min_keyword_score:
                normalized_keyword_score = (scores["keyword_score"] - min_keyword_score) / (max_keyword_score - min_keyword_score)
            else:
                normalized_keyword_score = 1.0 if scores["keyword_score"] > 0 else 0.0

            # 计算混合分数
            hybrid_score = (vector_weight * normalized_vector_score +
                           (1 - vector_weight) * normalized_keyword_score)

            # 创建结果对象
            result = scores["doc"].copy()
            result["score"] = hybrid_score
            # 添加详细分数，方便调试
            result["vector_score"] = scores["vector_score"]
            result["normalized_vector_score"] = normalized_vector_score
            result["keyword_score"] = scores["keyword_score"]
            result["normalized_keyword_score"] = normalized_keyword_score

            # 只添加有效分数的结果
            if hybrid_score > 0:
                results.append(result)

        # 如果没有有效结果，但原始检索有结果，则保留原始结果
        if not results and (vector_results or keyword_results):
            logger.warning("混合排名产生了零分数，使用原始检索结果")
            combined = vector_results + keyword_results
            # 去重
            seen_ids = set()
            results = []
            for r in combined:
                if r["id"] not in seen_ids:
                    seen_ids.add(r["id"])
                    r["score"] = r.get("score", 0.1)  # 确保有分数
                    results.append(r)

        # 按混合分数排序
        results.sort(key=lambda x: x["score"], reverse=True)

        # 更新排名
        for i, result in enumerate(results):
            result["rank"] = i

        return results

    def save_index(self, directory: str) -> None:
        """
        保存索引到指定目录

        Args:
            directory: 保存目录
        """
        if not os.path.exists(directory):
            os.makedirs(directory)

        # 保存FAISS索引
        faiss_path = os.path.join(directory, "faiss_index.bin")
        faiss.write_index(self.vector_index, faiss_path)

        # 保存文档
        docs_path = os.path.join(directory, "documents.json")
        with open(docs_path, 'w', encoding='utf-8') as f:
            json.dump(self.docs, f, ensure_ascii=False, indent=2)

        # 保存BM25索引（如果存在）
        if self.bm25_index:
            import pickle
            bm25_path = os.path.join(directory, "bm25_index.pkl")
            with open(bm25_path, 'wb') as f:
                pickle.dump({
                    "index": self.bm25_index,
                    "tokenized_corpus": self.tokenized_corpus
                }, f)

        logger.info(f"索引已保存到目录: {directory}")

    def load_index(self, directory: str) -> bool:
        """
        从指定目录加载索引

        Args:
            directory: 索引目录

        Returns:
            加载是否成功
        """
        try:
            # 加载FAISS索引
            faiss_path = os.path.join(directory, "faiss_index.bin")
            if os.path.exists(faiss_path):
                self.vector_index = faiss.read_index(faiss_path)
            else:
                logger.warning(f"FAISS索引文件不存在: {faiss_path}")
                return False

            # 加载文档
            docs_path = os.path.join(directory, "documents.json")
            if os.path.exists(docs_path):
                with open(docs_path, 'r', encoding='utf-8') as f:
                    self.docs = json.load(f)
            else:
                logger.warning(f"文档文件不存在: {docs_path}")
                return False

            # 加载BM25索引（如果存在）
            import pickle
            bm25_path = os.path.join(directory, "bm25_index.pkl")
            if os.path.exists(bm25_path) and self.bm25_class:
                with open(bm25_path, 'rb') as f:
                    bm25_data = pickle.load(f)
                    self.bm25_index = bm25_data["index"]
                    self.tokenized_corpus = bm25_data["tokenized_corpus"]

            self.index_built = True
            logger.info(f"索引加载成功，包含{len(self.docs)}个文档")
            return True

        except Exception as e:
            logger.error(f"加载索引失败: {str(e)}")
            return False


class DocumentSplitter(ABC):
    """文档切分基类"""

    @abstractmethod
    def split(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        将文档切分为多个片段

        Args:
            document: 包含text和metadata的文档字典

        Returns:
            切分后的文档片段列表
        """
        pass


class TextChunkSplitter(DocumentSplitter):
    """基于文本长度的文档切分器"""

    def __init__(self,
                 chunk_size: int = 500,
                 chunk_overlap: int = 50,
                 separator: str = "\n"):
        """
        初始化文本切分器

        Args:
            chunk_size: 每个文本块的最大字符数
            chunk_overlap: 相邻文本块的重叠字符数
            separator: 首选的分隔符
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.separator = separator

    def split(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        将文档按指定大小切分

        Args:
            document: 包含text和metadata的文档字典

        Returns:
            切分后的文档片段列表
        """
        text = document.get("text", "")
        metadata = document.get("metadata", {})

        if not text:
            return []

        chunks = []
        start = 0

        # 如果文本长度小于chunk_size，直接返回
        if len(text) <= self.chunk_size:
            return [{"text": text, "metadata": metadata}]

        while start < len(text):
            # 计算当前块的结束位置
            end = start + self.chunk_size

            # 如果已经到达文本末尾，直接截取到末尾
            if end >= len(text):
                chunks.append({"text": text[start:], "metadata": metadata.copy()})
                break

            # 否则，找到更好的分割点
            # 优先在分隔符处分割
            split_point = text.rfind(self.separator, start, end)

            if split_point != -1 and split_point > start:
                # 找到了合适的分隔符
                chunks.append({"text": text[start:split_point], "metadata": metadata.copy()})
                # 下一块的开始位置，考虑重叠
                start = max(split_point, start + self.chunk_size - self.chunk_overlap)
            else:
                # 没找到合适的分隔符，直接按长度切分
                chunks.append({"text": text[start:end], "metadata": metadata.copy()})
                # 下一块的开始位置，考虑重叠
                start = max(end, start + self.chunk_size - self.chunk_overlap)

        # 为每个块添加索引信息
        for i, chunk in enumerate(chunks):
            chunk["metadata"]["chunk_index"] = i
            chunk["metadata"]["chunk_count"] = len(chunks)

        return chunks


class DocumentLoader(ABC):
    """文档加载器基类"""

    @abstractmethod
    def load(self, file_path: str) -> Dict[str, Any]:
        """
        加载文档

        Args:
            file_path: 文档路径

        Returns:
            包含text和metadata的文档字典
        """
        pass


class PDFLoader(DocumentLoader):
    """PDF文档加载器"""

    def __init__(self, extraction_method: str = "pymupdf"):
        """
        初始化PDF加载器

        Args:
            extraction_method: 文本提取方法，支持"pymupdf"和"tesseract"
        """
        self.extraction_method = extraction_method

    def load(self, file_path: str) -> Dict[str, Any]:
        """
        加载PDF文档

        Args:
            file_path: PDF文件路径

        Returns:
            包含text和metadata的文档字典
        """
        try:
            file_name = os.path.basename(file_path)

            # 根据指定方法提取文本
            if self.extraction_method == "pymupdf":
                text = extract_text_with_pymupdf(file_path)
            elif self.extraction_method == "tesseract":
                text = extract_text_with_pytesseract(file_path)
            else:
                # 默认使用pymupdf
                text = extract_text_with_pymupdf(file_path)

            # 提取竞赛信息
            competition_info = extract_competition_info(text, file_name)

            # 构建元数据
            metadata = {
                "source": file_path,
                "file_name": file_name,
                "extraction_method": self.extraction_method,
                "creation_time": time.time(),
                **competition_info
            }

            return {
                "text": text,
                "metadata": metadata
            }

        except Exception as e:
            logger.error(f"加载PDF文件失败 {file_path}: {str(e)}")
            # 返回空文档而不是None，以便后续处理
            return {
                "text": "",
                "metadata": {
                    "source": file_path,
                    "file_name": os.path.basename(file_path),
                    "error": str(e)
                }
            }


class DocumentProcessingPipeline:
    """文档处理管道，处理从加载到拆分和索引构建的全流程"""

    def __init__(self,
                 loader: DocumentLoader = None,
                 splitter: DocumentSplitter = None,
                 retrieval_system: RetrievalSystem = None,
                 data_dir: str = "",
                 chunk_size: int = 800,
                 chunk_overlap: int = 200,
                 model_name: str = "shibing624/text2vec-base-chinese",
                 output_dir: str = "knowledge_base",
                 dense_weight: float = 0.7):
        """
        初始化文档处理管道

        Args:
            loader: 文档加载器，如果为None则创建默认的PDF加载器
            splitter: 文本拆分器，如果为None则创建默认的文本拆分器
            retrieval_system: 检索系统，如果为None则创建默认的检索系统
            data_dir: 文本文件所在目录
            chunk_size: 文本块大小
            chunk_overlap: 文本块重叠大小
            model_name: 向量模型名称
            output_dir: 输出目录
            dense_weight: 稠密检索权重
        """
        self.data_dir = data_dir
        self.output_dir = output_dir

        # 如果没有提供加载器，创建默认的PDF加载器
        self.loader = loader or PDFLoader()

        # 如果没有提供分割器，创建结构感知的文本分割器
        self.splitter = splitter or StructureAwareTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap
        )

        # 如果没有提供检索系统，创建默认的检索系统
        self.retrieval_system = retrieval_system or RetrievalSystem(
            vector_model_name=model_name,
            persist_dir=output_dir
        )

        # 加载文本处理器
        if data_dir:
            self.text_processor = TextProcessor(data_dir)

    @classmethod
    def load(cls, directory_path: str, model_name: str = "shibing624/text2vec-base-chinese") -> "DocumentProcessingPipeline":
        """
        从指定目录加载文档处理管道

        Args:
            directory_path: 索引目录路径
            model_name: 向量模型名称

        Returns:
            文档处理管道实例
        """
        # 创建检索系统
        retrieval_system = RetrievalSystem(
            vector_model_name=model_name,
            persist_dir=directory_path
        )

        # 创建和返回管道
        return cls(
            retrieval_system=retrieval_system,
            output_dir=directory_path
        )

    def process_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        处理单个文件

        Args:
            file_path: 文件路径

        Returns:
            处理后的文档块列表
        """
        # 1. 加载文档
        logger.info(f"加载文件: {file_path}")
        doc = self.loader.load(file_path)

        if not doc["text"]:
            logger.warning(f"文件无内容或加载失败: {file_path}")
            return []

        # 2. 拆分文档
        logger.info(f"拆分文档")
        chunks = self.splitter.split_documents([doc])

        logger.info(f"文档'{os.path.basename(file_path)}'已拆分为{len(chunks)}个块")
        return chunks

    def process_directory(self, directory_path: str, file_extension: str = ".pdf") -> List[Dict[str, Any]]:
        """
        处理目录中的所有指定类型文件

        Args:
            directory_path: 目录路径
            file_extension: 文件扩展名，默认为.pdf

        Returns:
            处理后的所有文档块列表
        """
        all_chunks = []

        # 递归查找所有匹配的文件
        for root, _, files in os.walk(directory_path):
            for file in files:
                if file.lower().endswith(file_extension.lower()):
                    file_path = os.path.join(root, file)
                    chunks = self.process_file(file_path)
                    all_chunks.extend(chunks)

        logger.info(f"共处理{len(all_chunks)}个文档块")
        return all_chunks

    def process_all_documents(self) -> bool:
        """
        处理所有文档并构建索引

        Returns:
            是否成功
        """
        try:
            all_chunks = []

            # 处理文本文件
            if hasattr(self, 'text_processor'):
                # 处理TXT文件
                for file_path in self.text_processor.txt_file_paths:
                    logger.info(f"处理文本文件: {file_path}")
                    text = self.text_processor.read_text_file(file_path)
                    metadata = self.text_processor.get_document_metadata(file_path)

                    # 创建文档
                    doc = {"text": text, "metadata": metadata}

                    # 切分文档
                    chunks = self.splitter.split(doc)
                    all_chunks.extend(chunks)

                # 处理Excel文件
                for file_path in self.text_processor.excel_file_paths:
                    excel_chunks = self.process_excel_file(file_path)
                    all_chunks.extend(excel_chunks)

            # 构建索引
            if all_chunks:
                logger.info(f"开始构建索引，共有{len(all_chunks)}个文档块")
                self.retrieval_system.build_index(all_chunks)

                # 保存配置信息
                config = {
                    "num_txt_files": len(getattr(self.text_processor, 'txt_file_paths', [])),
                    "num_excel_files": len(getattr(self.text_processor, 'excel_file_paths', [])),
                    "num_chunks": len(all_chunks),
                    "vector_dim": self.retrieval_system.vector_dim,
                    "dense_weight": getattr(self.retrieval_system, 'dense_weight', 0.7)
                }

                # 保存配置
                config_path = os.path.join(self.output_dir, "config.json")
                with open(config_path, 'w', encoding='utf-8') as f:
                    json.dump(config, f, indent=2, ensure_ascii=False)

                logger.info(f"索引构建完成，配置保存到{config_path}")
                return True
            else:
                logger.warning("没有文档可索引")
                return False

        except Exception as e:
            logger.error(f"处理文档时出错: {str(e)}")
            return False

    def process_excel_file(self, file_path: str) -> List[Dict[str, Any]]:
        """
        处理Excel文件

        Args:
            file_path: Excel文件路径

        Returns:
            处理后的文档块列表
        """
        try:
            # 使用TextProcessor读取Excel
            if hasattr(self, 'text_processor'):
                docs = self.text_processor.read_excel_file(file_path)
            else:
                # 如果没有TextProcessor，创建一个临时的来处理Excel
                temp_processor = TextProcessor("")
                docs = temp_processor.read_excel_file(file_path)

            # 切分每个文档
            all_chunks = []
            for doc in docs:
                chunks = self.splitter.split(doc)
                all_chunks.extend(chunks)

            logger.info(f"从Excel文件中处理了{len(all_chunks)}个文档块")
            return all_chunks

        except Exception as e:
            logger.error(f"处理Excel文件时出错: {str(e)}")
            return []

    def search_similar(self, query: str, top_k: int = 10, use_hybrid: bool = True) -> List[Dict[str, Any]]:
        """
        搜索相似文档

        Args:
            query: 查询文本
            top_k: 返回结果数量
            use_hybrid: 是否使用混合检索

        Returns:
            相似文档列表
        """
        # 调用检索系统搜索
        results = self.retrieval_system.search(
            query=query,
            top_k=top_k,
            hybrid_weight=0.7 if use_hybrid else 1.0
        )

        return results


class ZhipuAIClient:
    """智谱AI模型客户端"""

    def __init__(self, api_key: str, model: str = "glm-4"):
        """
        初始化智谱客户端

        Args:
            api_key: 智谱API密钥
            model: 模型名称，默认为glm-4
        """
        self.api_key = api_key
        self.model = model
        self.client = zhipuai.ZhipuAI(api_key=api_key)
        logger.info(f"初始化智谱AI客户端，使用模型: {model}")

    def generate(self, prompt: str, temperature: float = 0.7, max_tokens: int = 2000) -> str:
        """
        调用大模型生成回答

        Args:
            prompt: 提示词
            temperature: 温度参数，控制随机性
            max_tokens: 最大生成token数

        Returns:
            生成的回答文本
        """
        try:
            logger.info("正在调用智谱API生成回答...")
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "user", "content": prompt}
                ],
                temperature=temperature,
                max_tokens=max_tokens
            )

            if hasattr(response, 'choices') and len(response.choices) > 0:
                answer = response.choices[0].message.content
                logger.info(f"智谱API调用成功，生成了 {len(answer)} 字符的回答")
                return answer
            else:
                logger.error(f"智谱API返回了未预期的格式: {response}")
                return "抱歉，生成回答时发生错误。"

        except Exception as e:
            logger.error(f"调用智谱API时出错: {str(e)}")
            return f"抱歉，调用大模型API时发生错误: {str(e)}"

    def generate_with_rag(self,
                          query: str,
                          retrieved_docs: List[Dict[str, Any]],
                          temperature: float = 0.7) -> str:
        """
        基于检索到的文档生成回答

        Args:
            query: 用户查询
            retrieved_docs: 检索到的文档列表
            temperature: 温度参数

        Returns:
            生成的回答文本
        """
        # 构建RAG提示词
        prompt = self._build_rag_prompt(query, retrieved_docs)

        # 调用模型生成回答
        return self.generate(prompt, temperature)

    def _build_rag_prompt(self, query: str, retrieved_docs: List[Dict[str, Any]]) -> str:
        """
        构建RAG提示词

        Args:
            query: 用户查询
            retrieved_docs: 检索到的文档列表

        Returns:
            完整的提示词
        """
        # 提取文档内容
        contexts = []
        for i, doc in enumerate(retrieved_docs, 1):
            # 获取文档文本
            text = doc.get("text", "")

            # 获取文档来源
            metadata = doc.get("metadata", {})
            source = metadata.get("source", f"文档{i}")

            # 获取章节信息
            section = metadata.get("section", "")
            section_info = f"章节：{section}" if section else ""

            # 获取关键信息
            key_info = metadata.get("key_info", {})
            key_info_text = ""
            if key_info:
                key_info_parts = []
                for key, value in key_info.items():
                    key_info_parts.append(f"{key}: {value}")
                key_info_text = "关键信息：" + "；".join(key_info_parts)

            # 获取相关度分数
            score_info = ""
            if "score" in doc:
                score_info = f"(相关度: {doc['score']:.2f})"

            # 添加到上下文列表
            context_entry = f"[来源：{source}] {score_info}\n{section_info}\n{key_info_text}\n{text}"
            contexts.append(context_entry)

        # 合并上下文
        context_text = "\n\n".join(contexts)

        if context_text:
            prompt = f"""请根据提供的参考信息回答用户问题。你是一个专业的竞赛信息咨询顾问，请尽可能详细准确地回答用户的问题。如果无法从参考信息中找到答案，请明确说明无法回答，不要编造信息。
参考信息：
{context_text}

用户问题：{query}

请注意：
1. 回答要准确、具体且有针对性,并且尽量简洁不要一长串
2. 如果找到精确答案，请直接引用原文相关部分
3. 如有必要，可根据不同参考文档综合信息
4. 若多个文档提供了相同信息，以相关度更高的文档为准
5. 如果问题涉及报名时间、字数要求等具体数字信息，请确保准确引用

请现在回答用户问题："""
        else:
            prompt = f"""请回答用户的问题。如果无法回答，请明确告知无法回答，不要编造信息。            
用户问题：{query}
请给出准确、简洁的回答。"""

        return prompt


class RAGSystem:
    """检索增强生成系统"""

    def __init__(self,
                 knowledge_base_path: str,
                 api_key: str,
                 model_name: str = "glm-4",
                 embed_model_name: str = "all-MiniLM-L6-v2",
                 dense_weight: float = 0.7,
                 top_k: int = 10):
        """
        初始化RAG系统

        Args:
            knowledge_base_path: 知识库路径
            api_key: 智谱API密钥
            model_name: 智谱模型名称
            embed_model_name: 向量模型名称
            dense_weight: 稠密检索权重
            top_k: 检索文档数量
        """
        self.top_k = top_k

        # 加载知识库
        logger.info(f"加载知识库: {knowledge_base_path}")
        pipeline = DocumentProcessingPipeline.load(knowledge_base_path, embed_model_name)
        self.knowledge_base = pipeline.retrieval_system

        # 初始化智谱客户端
        self.llm_client = ZhipuAIClient(api_key, model_name)

        logger.info("RAG系统初始化完成")

    def answer_query(self,
                    query: str,
                    use_hybrid: bool = True,
                    temperature: float = 0.7,
                    history_manager: Optional['HistoryManager'] = None) -> Dict[str, Any]:
        """
        回答用户查询

        Args:
            query: 用户查询
            use_hybrid: 是否使用混合检索
            temperature: 温度参数
            history_manager: 历史记录管理器，如果提供则记录问答

        Returns:
            包含回答和检索结果的字典
        """
        start_time = time.time()

        # 步骤1: 检索相关文档
        logger.info(f"检索与问题相关的文档: {query}")
        retrieved_docs = self.knowledge_base.search(
            query,
            top_k=self.top_k,
            hybrid_weight=0.7 if use_hybrid else 1.0
        )

        retrieval_time = time.time() - start_time
        logger.info(f"检索完成，耗时 {retrieval_time:.2f} 秒，找到 {len(retrieved_docs)} 个相关文档")

        # 添加最终相关度分数用于显示
        for doc in retrieved_docs:
            doc["final_score"] = doc.get("score", 0)

        # 步骤2: 生成回答
        logger.info("生成回答...")
        answer = self.llm_client.generate_with_rag(
            query,
            retrieved_docs,
            temperature
        )

        total_time = time.time() - start_time
        logger.info(f"回答生成完成，总耗时 {total_time:.2f} 秒")

        # 步骤3: 如果提供了历史管理器，记录问答
        if history_manager is not None:
            history_manager.add_qa_record(query, answer, retrieved_docs)

        # 整理返回结果
        result = {
            "query": query,
            "answer": answer,
            "retrieved_documents": retrieved_docs,
            "metrics": {
                "retrieval_time": retrieval_time,
                "total_time": total_time,
                "num_docs_retrieved": len(retrieved_docs)
            }
        }

        return result


class KeywordEnhancer:
    """关键词增强器：扩展同义词和提取关键信息"""

    def __init__(self):
        """初始化关键词增强器"""
        # 常见竞赛关键词同义词表
        self.synonym_map = {
            # 竞赛类型同义词
            "专项赛": ["专项赛", "挑战赛", "竞赛", "大赛", "比赛"],
            "挑战赛": ["专项赛", "挑战赛", "竞赛", "大赛", "比赛"],
            "竞赛": ["专项赛", "挑战赛", "竞赛", "大赛", "比赛"],

            # 时间相关同义词
            "报名时间": ["报名时间", "报名日期", "报名截止", "注册时间", "注册期限"],
            "比赛时间": ["比赛时间", "竞赛时间", "赛程时间", "比赛日期", "竞赛日期"],

            # 要求相关同义词
            "参赛要求": ["参赛要求", "参赛条件", "参赛资格", "报名条件", "参与条件"],
            "参赛对象": ["参赛对象", "参赛人群", "参与对象", "面向人群", "适合人群"],
            "字数要求": ["字数要求", "字数限制", "文字数量", "字符数", "不少于", "不超过"],

            # 自动化控制相关同义词
            "自动化控制": ["自动化控制", "自动控制", "智能控制", "程序控制", "感应控制",
                     "自动化系统", "智能系统", "编写控制程序", "控制逻辑"],

            # 任务相关同义词
            "任务": ["任务", "挑战", "题目", "问题", "项目"],
            "智能交通": ["智能交通", "交通信号灯", "红绿灯", "交通系统", "信号灯控制"],
            "自动浇灌": ["自动浇灌", "智能浇灌", "浇水系统", "灌溉系统", "浇花"],
            "智能灯": ["智能灯", "感应灯", "光控灯", "声控灯", "人体感应"],
            "智能空调": ["智能空调", "温度控制", "空调系统", "温控系统"],
            "自动门": ["自动门", "感应门", "自动感应门", "门控系统"],

            # 准备工作相关同义词
            "准备工作": ["准备工作", "参赛准备", "比赛准备", "报名准备", "参赛流程",
                     "准备事项", "注意事项", "参赛步骤", "报名步骤", "报名流程",
                     "前期准备", "参赛要求", "报名要求", "须知", "申报要求"]
        }

        # 任务编号与标题映射
        self.task_map = {
            "任务一": ["智能交通信号灯", "交通信号灯", "红绿灯"],
            "任务二": ["智能灯", "光控灯", "声控灯", "人体感应灯"],
            "任务三": ["自动浇灌系统", "浇水系统", "灌溉系统"],
            "任务四": ["电子琴", "音乐演奏", "声音控制"],
            "任务五": ["智能空调", "温度控制", "空调系统"],
            "任务六": ["自动感应门", "感应门", "自动门"]
        }

        # 竞赛常见章节标题
        self.section_titles = [
            "竞赛简介", "比赛背景", "竞赛背景", "比赛目的", "竞赛目的",
            "参赛对象", "参赛要求", "参赛条件", "参赛资格",
            "报名方式", "报名时间", "比赛时间", "竞赛流程",
            "评分标准", "评分规则", "评审标准", "评审规则",
            "奖项设置", "奖励方式", "获奖名额",
            "参赛成果", "成果提交", "作品提交",
            "任务描述", "赛题介绍", "比赛任务"
        ]

        # 特殊竞赛处理映射
        self.special_competitions = {
            "未来校园智能应用专项赛": ["未来校园", "智能应用", "01_", "未来校园智能应用"],
            "智能数据采集装置设计专项赛": ["智能数据采集", "数据采集装置", "采集装置设计"],
            "太空探索智能机器人专项赛": ["太空探索", "智能机器人", "太空机器人"],
            "无人驾驶智能车专项赛": ["无人驾驶", "智能车", "无人车"]
        }

    def enhance_query(self, query: str) -> str:
        """
        增强查询，添加同义词和关键结构

        Args:
            query: 原始查询

        Returns:
            增强后的查询
        """
        original_query = query
        enhanced_parts = []

        # 1. 尝试提取竞赛名称
        competition_match = re.search(r'[""]?([^""]+?(?:专项赛|挑战赛|竞赛|大赛))[""]?', query)
        if competition_match:
            competition_name = competition_match.group(1)
            enhanced_parts.append(f"竞赛:{competition_name}")

            # 查找特殊竞赛处理
            clean_name = competition_name.replace('"', '').replace('"', '')
            for special_comp, alternatives in self.special_competitions.items():
                if special_comp in clean_name or any(alt in clean_name for alt in alternatives):
                    for alt in alternatives:
                        enhanced_parts.append(f"竞赛别名:{alt}")

        # 2. 检查是否包含任务编号
        for task_num, keywords in self.task_map.items():
            if task_num in query:
                enhanced_parts.append(f"任务编号:{task_num}")
                # 添加对应的任务关键词
                for keyword in keywords:
                    enhanced_parts.append(f"任务关键词:{keyword}")

        # 3. 检查是否包含章节标题关键词
        for title in self.section_titles:
            if title in query:
                enhanced_parts.append(f"章节:{title}")

        # 4. 检查是否包含"准备工作"相关关键词
        if any(term in query for term in ["准备", "准备工作", "参赛准备", "报名准备"]):
            enhanced_parts.append("主题:准备工作")
            enhanced_parts.append("章节:参赛要求")
            enhanced_parts.append("章节:报名方式")
            enhanced_parts.append("章节:比赛流程")
            enhanced_parts.append("章节:注意事项")

        # 5. 扩展同义词
        for key_term, synonyms in self.synonym_map.items():
            if key_term in query:
                # 添加同义词扩展
                syn_terms = " OR ".join([f'"{syn}"' for syn in synonyms if syn != key_term])
                if syn_terms:
                    enhanced_parts.append(f"同义词:({key_term} OR {syn_terms})")

        # 如果有增强部分，添加到原始查询
        if enhanced_parts:
            enhanced_query = f"{original_query} {' '.join(enhanced_parts)}"
            logger.info(f"查询增强: {original_query} -> {enhanced_query}")
            return enhanced_query

        # 没有找到可增强的部分，返回原始查询
        return original_query

    def extract_keywords(self, query: str) -> List[str]:
        """
        从查询中提取关键词

        Args:
            query: 查询文本

        Returns:
            关键词列表
        """
        keywords = []

        # 1. 提取竞赛名称
        competition_match = re.search(r'[""]?([^""]+(?:专项赛|挑战赛|竞赛|大赛))[""]?', query)
        if competition_match:
            competition_name = competition_match.group(1)
            keywords.append(competition_name)

        # 2. 提取任务编号和描述
        for task_num, task_keywords in self.task_map.items():
            if task_num in query:
                keywords.append(task_num)
                # 查找任务关键词
                for task_kw in task_keywords:
                    if task_kw in query:
                        keywords.append(task_kw)

        # 3. 提取同义词表中的关键词
        for key_term in self.synonym_map.keys():
            if key_term in query:
                keywords.append(key_term)

        # 4. 提取章节标题
        for title in self.section_titles:
            if title in query:
                keywords.append(title)

        # 5. 提取其他长度大于1的词语
        other_words = [w for w in jieba.lcut_for_search(query)
                      if len(w) > 1 and w not in keywords]
        keywords.extend(other_words)

        return keywords

    def get_synonyms(self, keyword: str) -> List[str]:
        """
        获取关键词的同义词

        Args:
            keyword: 关键词

        Returns:
            同义词列表
        """
        # 直接查找同义词表
        for key, synonyms in self.synonym_map.items():
            if keyword == key or keyword in synonyms:
                return synonyms

        # 检查任务编号
        for task_num, task_keywords in self.task_map.items():
            if keyword == task_num or keyword in task_keywords:
                result = [task_num] + task_keywords
                return list(set(result))  # 去重

        # 没有找到同义词
        return [keyword]


class StructureAwareTextSplitter(DocumentSplitter):
    """结构感知的文本切分器：保留文档结构信息"""

    def __init__(self,
                 chunk_size: int = 500,
                 chunk_overlap: int = 100,
                 separators: List[str] = None,
                 keep_section_titles: bool = True):
        """
        初始化结构感知的文本切分器

        Args:
            chunk_size: 文本块的最大字符数
            chunk_overlap: 文本块的重叠字符数
            separators: 分隔符列表，按优先级排序
            keep_section_titles: 是否保留章节标题
        """
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap

        # 默认分隔符
        if separators is None:
            self.separators = [
                "\n\n",  # 段落分隔
                "\n",    # 换行
                "。",    # 句号
                "！",    # 感叹号
                "？",    # 问号
                "；",    # 分号
                "，",    # 逗号
                " ",     # 空格
                ""       # 字符级别
            ]
        else:
            self.separators = separators

        self.keep_section_titles = keep_section_titles

        # 章节标题模式
        self.section_patterns = [
            # 数字编号标题
            r"^\d+[\.\s、]+\s*([^\n]+)$",
            # 中文数字编号标题
            r"^[一二三四五六七八九十]+[、\s]+\s*([^\n]+)$",
            # 带括号的编号标题
            r"^\([\d一二三四五六七八九十]+\)\s*([^\n]+)$",
            # 任务标题
            r"^任务[一二三四五六七八九十\d]+[：:]?\s*([^\n]+)$",
            # 常见章节标题
            r"^(参赛要求|参赛对象|报名方式|报名时间|比赛时间|奖项设置|评分标准|任务描述|竞赛简介|比赛背景)[：:]\s*([^\n]*)$"
        ]

        # 关键信息模式
        self.key_info_patterns = [
            # 报名时间
            r"报名时间[：:]\s*(.+?)(?:\n|$|。)",
            # 比赛时间
            r"(?:比赛|竞赛)时间[：:]\s*(.+?)(?:\n|$|。)",
            # 字数要求
            r"(?:字数要求|字数限制)[：:]\s*(.+?)(?:\n|$|。)",
            r"(?:不少于|不超过)\s*(\d+)\s*字",
            # 联系方式
            r"(?:联系方式|联系电话|咨询电话)[：:]\s*(.+?)(?:\n|$|。)"
        ]

    def _extract_section_titles(self, text: str) -> List[Tuple[int, str]]:
        """
        提取文本中的章节标题及其位置

        Args:
            text: 输入文本

        Returns:
            包含(位置, 标题)的列表
        """
        section_titles = []

        # 按行处理
        lines = text.split("\n")

        offset = 0
        for line in lines:
            stripped_line = line.strip()

            if stripped_line:
                # 尝试匹配所有章节标题模式
                for pattern in self.section_patterns:
                    match = re.match(pattern, stripped_line)
                    if match:
                        if len(match.groups()) == 1:
                            title = match.group(1)
                        else:
                            title = stripped_line

                        # 保存位置和标题
                        section_titles.append((offset, title))
                        break

            offset += len(line) + 1  # +1 for the newline

        return section_titles

    def _extract_key_info(self, text: str, extract_patterns: List[str] = None) -> Dict[str, str]:
        """
        提取文本中的关键信息

        Args:
            text: 输入文本
            extract_patterns: 自定义提取模式

        Returns:
            关键信息字典
        """
        key_info = {}

        patterns = extract_patterns or self.key_info_patterns

        for pattern in patterns:
            matches = re.finditer(pattern, text)
            for match in matches:
                # 提取关键信息
                if len(match.groups()) > 0:
                    info_type = pattern.split("[：:]")[0] if "[：:]" in pattern else "关键信息"
                    info_value = match.group(1)
                    key_info[info_type] = info_value

        return key_info

    def split(self, document: Dict[str, Any]) -> List[Dict[str, Any]]:
        """
        将文档切分为多个块，保留结构信息

        Args:
            document: 输入文档字典，包含text和metadata

        Returns:
            切分后的文档块列表
        """
        text = document.get("text", "")
        metadata = document.get("metadata", {}).copy()

        if not text:
            return []

        # 1. 提取章节标题和位置
        section_titles = self._extract_section_titles(text)

        # 2. 提取关键信息
        key_info = self._extract_key_info(text)
        if key_info:
            # 将关键信息添加到元数据
            metadata["key_info"] = key_info

        # 3. 准备分块
        chunks = []

        # 如果文本长度小于chunk_size，直接返回
        if len(text) <= self.chunk_size:
            return [{"text": text, "metadata": metadata}]

        # 4. 分块处理
        start = 0
        current_section = None

        while start < len(text):
            # 确定当前位置所在的章节
            for pos, title in section_titles:
                if pos <= start and (current_section is None or pos > current_section[0]):
                    current_section = (pos, title)

            # 计算当前块的结束位置
            end = start + self.chunk_size

            # 检查是否已到文本末尾
            if end >= len(text):
                # 创建最后一个文本块
                chunk_text = text[start:]
                chunk_metadata = metadata.copy()

                # 添加章节信息
                if current_section and self.keep_section_titles:
                    chunk_metadata["section"] = current_section[1]

                chunks.append({"text": chunk_text, "metadata": chunk_metadata})
                break

            # 寻找最佳切分点
            split_point = None

            # 按优先级尝试不同的分隔符
            for sep in self.separators:
                # 从结束位置向前查找分隔符
                candidate = text.rfind(sep, start, end)
                if candidate > start:  # 找到有效的分隔符位置
                    split_point = candidate + len(sep)
                    break

            # 如果找不到合适的分隔符，直接在end处切分
            if split_point is None or split_point <= start:
                split_point = end

            # 创建当前块
            chunk_text = text[start:split_point]
            chunk_metadata = metadata.copy()

            # 添加章节信息
            if current_section and self.keep_section_titles:
                chunk_metadata["section"] = current_section[1]

                # 如果当前块不包含章节标题，在文本开头添加
                if current_section[0] < start and self.keep_section_titles:
                    section_prefix = f"【{current_section[1]}】\n"
                    chunk_text = section_prefix + chunk_text

            # 添加到结果列表
            chunks.append({"text": chunk_text, "metadata": chunk_metadata})

            # 更新下一块的起始位置，考虑重叠
            start = max(split_point - self.chunk_overlap, start + 1)

        # 5. 为每个块添加索引信息
        for i, chunk in enumerate(chunks):
            chunk["metadata"]["chunk_index"] = i
            chunk["metadata"]["chunk_count"] = len(chunks)

        return chunks


class HistoryManager:
    """历史问答记录管理器：管理问答历史并支持导出到Excel"""

    def __init__(self, history_file: str = "qa_history.json"):
        """
        初始化历史记录管理器

        Args:
            history_file: 历史记录JSON文件路径
        """
        self.history_file = history_file
        self.history = self._load_history()

    def _load_history(self) -> List[Dict[str, Any]]:
        """从文件加载历史记录"""
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"加载历史记录失败: {str(e)}")
                return []
        return []

    def _save_history(self) -> None:
        """保存历史记录到文件"""
        try:
            logger.info(f"正在保存历史记录到文件 {self.history_file}，共 {len(self.history)} 条记录")
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.history, f, ensure_ascii=False, indent=2)
            logger.info(f"历史记录保存成功: {self.history_file}")
        except Exception as e:
            logger.error(f"保存历史记录失败: {str(e)}")
            # 尝试保存到当前目录
            try:
                alternative_file = f"qa_history_{time.strftime('%Y%m%d%H%M%S')}.json"
                logger.info(f"尝试保存到替代文件: {alternative_file}")
                with open(alternative_file, 'w', encoding='utf-8') as f:
                    json.dump(self.history, f, ensure_ascii=False, indent=2)
                logger.info(f"历史记录已保存到替代文件: {alternative_file}")
            except Exception as e2:
                logger.error(f"保存到替代文件也失败: {str(e2)}")

    def add_qa_record(self, query: str, answer: str, retrieved_docs: List[Dict[str, Any]]) -> None:
        """
        添加问答记录

        Args:
            query: 用户查询
            answer: 系统回答
            retrieved_docs: 检索到的文档
        """
        # 提取关键点
        key_points = self._extract_key_points(answer, retrieved_docs)

        record = {
            "id": f"C{len(self.history) + 1:03d}",
            "query": query,
            "answer": answer,
            "key_points": key_points,
            "timestamp": time.strftime("%Y-%m-%d %H:%M:%S"),
            "retrieved_docs": [
                {
                    "source": doc.get("metadata", {}).get("source", "未知"),
                    "score": doc.get("score", 0)
                }
                for doc in retrieved_docs[:3]  # 只保存前3个检索结果
            ]
        }

        self.history.append(record)
        self._save_history()
        logger.info(f"已添加问答记录 {record['id']}")

    def _extract_key_points(self, answer: str, docs: List[Dict[str, Any]]) -> str:
        """
        提取答案关键点

        通过智谱API生成答案摘要，提取关键信息点。

        Args:
            answer: 回答文本
            docs: 检索到的文档

        Returns:
            提取的关键点字符串
        """
        try:
            # 如果答案为空，返回空字符串
            if not answer or len(answer) < 10:
                return ""

            # 初始化智谱客户端
            api_key = "c66be1dbcd07484fa81efde6d883e410.O2y7leHySzSP7Ygc"
            llm_client = ZhipuAIClient(api_key, model="glm-4")
            logger.info("使用智谱API提取答案关键点")

            # 构建提取摘要的提示词
            prompt = f"""请从以下回答中提取1-3个关键信息点，以简明扼要的方式概括核心内容。
            每个要点应该是信息量较大的短句，不超过30字，并使用分号分隔。
            不要使用编号或序号，直接给出要点即可。
回答内容：
{answer}

要求:
0. 语句要通顺！！！！
1. 只提取回答中已有的关键信息，不要添加新信息
2. 关注数字、日期、截止时间、条件等具体细节
3. 保持客观，不要添加评价或建议
4. 直接列出关键点，不要有任何开头语（如"关键点如下"）
5. 总共不超过100字

关键信息点:"""

            # 调用API获取关键点
            key_points = llm_client.generate(prompt, temperature=0.3, max_tokens=300)

            # 清理结果
            key_points = key_points.strip()

            # 如果结果为空或出错，使用备用方法
            if not key_points or "抱歉" in key_points or "错误" in key_points:
                logger.warning("智谱API提取关键点失败，使用备用方法")
                # 截取前150个字符作为备用
                return answer[:150] + "..." if len(answer) > 150 else answer

            logger.info(f"成功提取 {len(key_points)} 字符的关键点")
            return key_points

        except Exception as e:
            logger.error(f"提取答案关键点时出错: {str(e)}")
            # 发生错误时使用答案的前150个字符
            return answer[:150] + "..." if len(answer) > 150 else answer

    def export_to_excel(self, excel_file: str = "qa_history.xlsx") -> str:
        """
        导出历史记录到Excel文件

        Args:
            excel_file: 导出的Excel文件路径

        Returns:
            Excel文件路径
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill

            # 准备数据
            data = []
            for record in self.history:
                data.append({
                    "问题编号": record["id"],
                    "问题": record["query"],
                    "关键点": record.get("key_points", ""),
                    "回答": record["answer"]
                })

            # 创建DataFrame
            df = pd.DataFrame(data)

            # 保存到Excel
            df.to_excel(excel_file, index=False, engine='openpyxl')
            logger.info(f"创建初始Excel文件: {excel_file}")

            # 使用openpyxl美化Excel
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active

            # 设置列宽
            ws.column_dimensions['A'].width = 10  # 问题编号
            ws.column_dimensions['B'].width = 40  # 问题
            ws.column_dimensions['C'].width = 50  # 关键点
            ws.column_dimensions['D'].width = 60  # 回答

            # 设置标题行样式
            header_fill = PatternFill(start_color="FFCCCCFF", end_color="FFCCCCFF", fill_type="solid")
            header_font = Font(bold=True, size=12)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 设置所有单元格的自动换行
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # 保存美化后的Excel
            wb.save(excel_file)

            logger.info(f"历史记录已导出到Excel文件: {excel_file}")
            return excel_file

        except ImportError as e:
            logger.error(f"导出Excel失败，缺少依赖包: {str(e)}")
            logger.info("请安装必要的依赖: pip install pandas openpyxl")
            return ""
        except Exception as e:
            logger.error(f"导出Excel失败: {str(e)}")
            return ""


def main():
    """主函数"""
    import argparse

    # 创建参数解析器
    parser = argparse.ArgumentParser(description="混合检索知识库构建工具")

    # 基础参数
    parser.add_argument("--data_dir", default="txts", help="文本文件目录")
    parser.add_argument("--output_dir", default="knowledge_base", help="知识库目录")
    parser.add_argument("--chunk_size", type=int, default=800, help="分块大小")
    parser.add_argument("--chunk_overlap", type=int, default=200, help="分块重叠大小")

    # 向量嵌入参数
    parser.add_argument("--model_name", default="all-MiniLM-L6-v2", help="向量化模型名称")
    parser.add_argument("--dense_weight", type=float, default=0.7, help="稠密检索权重(0-1之间)")

    # RAG参数
    parser.add_argument("--api_key", default="", help="智谱API密钥")
    parser.add_argument("--zhipu_model", default="glm-4", help="智谱模型名称")
    parser.add_argument("--query", default="", help="执行查询")
    parser.add_argument("--query_file", default="", help="从文件读取查询内容")

    parser.add_argument("--temperature", type=float, default=0.7, help="生成温度")
    parser.add_argument("--top_k", type=int, default=10, help="检索文档数量")

    # 操作模式
    parser.add_argument("--build_kb", action="store_true", help="构建知识库")
    parser.add_argument("--retrieve_only", action="store_true", help="仅执行检索，不生成回答")
    parser.add_argument("--no_hybrid", action="store_true", help="不使用混合检索")

    # 添加Excel相关参数
    parser.add_argument("--excel_file", default="result_1.xlsx", help="要处理的Excel文件路径")

    # 添加历史记录相关参数
    parser.add_argument("--save_history", action="store_true", help="保存问答历史记录")
    parser.add_argument("--history_file", default="qa_history.json", help="历史记录JSON文件路径")
    parser.add_argument("--export_excel", default="", help="导出历史记录到Excel文件")
    parser.add_argument("--batch_queries", default="", help="批量查询文件，每行一个查询")

    args = parser.parse_args()

    # 如果没有提供API密钥，尝试从环境变量获取
    api_key = args.api_key or os.environ.get("ZHIPU_API_KEY") or "c66be1dbcd07484fa81efde6d883e410.O2y7leHySzSP7Ygc"

    # 获取查询内容
    query = args.query

    # 如果提供了查询文件，则从文件读取查询内容
    if args.query_file and os.path.exists(args.query_file):
        try:
            with open(args.query_file, 'r', encoding='utf-8') as f:
                query = f.read().strip()
            logger.info(f"已从文件 {args.query_file} 读取查询内容")
        except Exception as e:
            logger.error(f"读取查询文件时出错: {e}")


    # 初始化历史记录管理器（如果需要）
    history_manager = None
    if args.save_history or args.export_excel:
        history_manager = HistoryManager(args.history_file)
        logger.info(f"初始化历史记录管理器，使用文件: {args.history_file}")

    # 如果请求导出Excel文件
    if args.export_excel:
        excel_path = history_manager.export_to_excel(args.export_excel)
        if excel_path:
            logger.info(f"历史记录已导出到: {excel_path}")
        else:
            logger.error("导出历史记录失败")
        # 如果只是导出Excel，不执行其他操作，则返回
        if not query and not args.build_kb and not args.process_excel_only and not args.batch_queries:
            return


    # 模式1: 构建知识库
    elif args.build_kb:
        logger.info(f"从 {args.data_dir} 构建知识库")
        pipeline = DocumentProcessingPipeline(
            data_dir=args.data_dir,
            chunk_size=args.chunk_size,
            chunk_overlap=args.chunk_overlap,
            model_name=args.model_name,
            output_dir=args.output_dir,
            dense_weight=args.dense_weight
        )

        # 处理所有文档
        pipeline.process_all_documents()
        logger.info(f"知识库构建完成，保存到 {args.output_dir}")

    # 模式2: 执行查询
    elif query or args.batch_queries:
        # 批量查询处理
        if args.batch_queries and os.path.exists(args.batch_queries):
            logger.info(f"执行批量查询: {args.batch_queries}")

            # 初始化RAG系统
            rag_system = RAGSystem(
                knowledge_base_path=args.output_dir,
                api_key=api_key,
                model_name=args.zhipu_model,
                embed_model_name=args.model_name,
                dense_weight=args.dense_weight,
                top_k=args.top_k
            )

            # 读取批量查询文件
            batch_queries = []
            try:
                with open(args.batch_queries, 'r', encoding='utf-8') as f:
                    batch_queries = [line.strip() for line in f if line.strip()]
                logger.info(f"加载了 {len(batch_queries)} 个查询")
            except Exception as e:
                logger.error(f"读取批量查询文件出错: {e}")
                return

            # 处理每个查询
            for i, query in enumerate(batch_queries, 1):
                logger.info(f"处理查询 {i}/{len(batch_queries)}: {query}")

                # 处理查询
                result = rag_system.answer_query(
                    query,
                    use_hybrid=not args.no_hybrid,
                    temperature=args.temperature,
                    history_manager=history_manager
                )

                # 输出回答
                logger.info(f"回答: {result['answer'][:100]}...")

            # 导出历史记录
            if history_manager and not args.export_excel:
                excel_path = history_manager.export_to_excel(f"result_2.xlsx")#这里直接要求改为result2.xlsx
                logger.info(f"批量查询结果已导出到: {excel_path}")

            return

        # 如果只需要检索
        if args.retrieve_only:
            logger.info("仅执行检索模式")
            # 加载知识库
            pipeline = DocumentProcessingPipeline.load(args.output_dir, args.model_name)

            # 执行检索
            results = pipeline.search_similar(query, top_k=args.top_k, use_hybrid=not args.no_hybrid)

            # 显示检索结果
            logger.info(f"检索结果，共 {len(results)} 个相关文档:")
            for i, result in enumerate(results, 1):
                # 显示检索分数
                score_info = ""
                if "vector_score" in result:
                    score_info = f"(向量:{result['vector_score']:.4f}, BM25:{result['keyword_score']:.4f}, 综合:{result['score']:.4f})"
                else:
                    score_info = f"(距离:{result.get('score', 0):.4f})"

                logger.info(f"{i}. 来源: {result['metadata']['source']} {score_info}")
                logger.info(f"   内容: {result['text'][:200]}...")
                logger.info("")

        # RAG模式: 检索 + 生成
        else:
            logger.info("RAG模式: 检索 + 生成")
            # 初始化RAG系统
            rag_system = RAGSystem(
                knowledge_base_path=args.output_dir,
                api_key=api_key,
                model_name=args.zhipu_model,
                embed_model_name=args.model_name,
                dense_weight=args.dense_weight,
                top_k=args.top_k
            )

            # 保存查询到文件以便调试
            query_debug_file = os.path.join(args.output_dir, "last_query.txt")
            try:
                with open(query_debug_file, 'w', encoding='utf-8') as f:
                    f.write(query)
            except Exception as e:
                logger.warning(f"保存查询到调试文件时出错: {e}")

            # 处理查询
            result = rag_system.answer_query(
                query,
                use_hybrid=not args.no_hybrid,
                temperature=args.temperature,
                history_manager=history_manager
            )

            # 输出回答
            logger.info("\n" + "="*80)
            logger.info(f"问题: {result['query']}")
            logger.info("-"*80)
            logger.info(f"回答: {result['answer']}")
            logger.info("="*80 + "\n")

            # 保存回答到文件
            answer_file = os.path.join(args.output_dir, "last_answer.txt")
            try:
                with open(answer_file, 'w', encoding='utf-8') as f:
                    f.write(f"问题: {result['query']}\n\n")
                    f.write(f"回答: {result['answer']}\n\n")
                    f.write("-"*80 + "\n")
                    f.write(f"检索到 {len(result['retrieved_documents'])} 个相关文档:\n")
                    for i, doc in enumerate(result['retrieved_documents'], 1):
                        source = doc['metadata']['source']
                        score = doc.get('score', doc.get('final_score', 0))
                        f.write(f"{i}. {source} (相关度: {score:.4f})\n")
                logger.info(f"结果已保存到 {answer_file}")
            except Exception as e:
                logger.warning(f"保存回答到文件时出错: {e}")

            # 输出检索文档信息
            logger.info(f"检索到 {len(result['retrieved_documents'])} 个相关文档:")
            for i, doc in enumerate(result['retrieved_documents'], 1):
                source = doc['metadata']['source']
                score = doc.get('score', doc.get('final_score', 0))
                logger.info(f"{i}. {source} (相关度: {score:.4f})")

            # 输出性能指标
            logger.info("\n性能指标:")
            for key, value in result['metrics'].items():
                if isinstance(value, float):
                    logger.info(f"- {key}: {value:.2f}")
                else:
                    logger.info(f"- {key}: {value}")

    else:
        parser.print_help()


if __name__ == "__main__":
    main()
